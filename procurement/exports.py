# procurement/exports.py
import io
from datetime import datetime

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.views import View
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from xhtml2pdf import pisa

from core.mixins import ModulePermissionMixin
from .models import RawMaterialReception, RawMaterialReceptionLine, PurchaseOrder, PurchaseOrderLine


# ---------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------

def _thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=10),
        "fill": PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _thin_border(),
    }


def _cell_style():
    return {
        "font": Font(size=10),
        "alignment": Alignment(vertical="center", wrap_text=True),
        "border": _thin_border(),
    }


def _apply_styles(cell, styles):
    for attr, value in styles.items():
        setattr(cell, attr, value)


def _logo_path():
    return str(settings.BASE_DIR / "static" / "img" / "logo.png")


# ---------------------------------------------------------------
# PDF — Listado general de recepciones
# ---------------------------------------------------------------

class ReceptionListPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "procurement.view_rawmaterialreception"

    def get(self, request):
        receptions = (
            RawMaterialReception.objects
            .select_related("purchase_order")
            .order_by("-reception_date", "-id")
        )
        html_string = render_to_string(
            "procurement/pdf_reception_list.html",
            {"receptions": receptions, "now": datetime.now(), "logo": _logo_path()},
        )
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="recepciones_mp.pdf"'
        pisa.CreatePDF(io.StringIO(html_string), dest=response)
        return response


# ---------------------------------------------------------------
# Excel — Listado general de recepciones
# ---------------------------------------------------------------

class ReceptionListExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "procurement.view_rawmaterialreception"

    def get(self, request):
        receptions = (
            RawMaterialReception.objects
            .select_related("purchase_order")
            .order_by("-reception_date", "-id")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Recepciones MP"

        headers = ["Código", "Fecha", "Proveedor", "RUC", "Documento", "N° Doc.", "Estado", "Cajas", "Peso Bruto", "Peso Neto", "Temp. °C"]
        hs = _header_style()
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            _apply_styles(cell, hs)

        cs = _cell_style()
        for row, r in enumerate(receptions, 2):
            values = [
                r.code,
                r.reception_date.strftime("%d/%m/%Y") if r.reception_date else "",
                r.supplier_name,
                r.supplier_ruc or "",
                r.document_type or "",
                r.document_number or "",
                r.get_status_display(),
                r.num_boxes,
                float(r.gross_weight) if r.gross_weight else None,
                float(r.net_weight) if r.net_weight else None,
                float(r.temperature_recorded) if r.temperature_recorded else None,
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_styles(cell, cs)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="recepciones_mp.xlsx"'
        return response


# ---------------------------------------------------------------
# PDF — Detalle de una recepción
# ---------------------------------------------------------------

class ReceptionDetailPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "procurement.view_rawmaterialreception"

    def get(self, request, pk):
        reception = get_object_or_404(
            RawMaterialReception.objects.select_related("purchase_order"), pk=pk
        )
        lines = reception.lines.select_related("product", "unit").order_by("id")
        html_string = render_to_string(
            "procurement/pdf_reception_detail.html",
            {"r": reception, "lines": lines, "now": datetime.now(), "logo": _logo_path()},
        )
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="recepcion_{reception.code}.pdf"'
        pisa.CreatePDF(io.StringIO(html_string), dest=response)
        return response


# ---------------------------------------------------------------
# Excel — Detalle de una recepción
# ---------------------------------------------------------------

class ReceptionDetailExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "procurement.view_rawmaterialreception"

    def get(self, request, pk):
        reception = get_object_or_404(
            RawMaterialReception.objects.select_related("purchase_order"), pk=pk
        )
        lines = reception.lines.select_related("product", "unit").order_by("id")

        wb = Workbook()
        ws = wb.active
        ws.title = "Recepción"

        # — Cabecera —
        title_font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")
        ws["A1"].value = f"Recepción {reception.code}"
        ws["A1"].font = title_font

        info = [
            ("Fecha:", reception.reception_date.strftime("%d/%m/%Y") if reception.reception_date else ""),
            ("Proveedor:", reception.supplier_name),
            ("RUC:", reception.supplier_ruc or ""),
            ("Documento:", f"{reception.document_type or ''} {reception.document_number or ''}".strip()),
            ("Estado:", reception.get_status_display()),
            ("OC:", reception.purchase_order.number if reception.purchase_order else "—"),
            ("Transporte:", f"{reception.transport_company or ''} / {reception.transport_plate or ''}".strip(" /")),
            ("Temp. °C:", str(reception.temperature_recorded) if reception.temperature_recorded else ""),
            ("Cajas:", str(reception.num_boxes) if reception.num_boxes else ""),
            ("Peso bruto:", str(reception.gross_weight) if reception.gross_weight else ""),
            ("Peso neto:", str(reception.net_weight) if reception.net_weight else ""),
        ]
        label_font = Font(bold=True, size=10)
        row = 3
        for label, value in info:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            row += 1

        # — Líneas —
        row += 1
        ws.cell(row=row, column=1, value="Detalle de líneas").font = Font(bold=True, size=12)
        row += 1

        line_headers = ["Producto", "Número de Lote", "Cant. Esperada", "Cant. Recibida", "Unidad", "Costo Unit.", "Vencimiento"]
        hs = _header_style()
        for col, h in enumerate(line_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            _apply_styles(cell, hs)

        cs = _cell_style()
        for line in lines:
            row += 1
            values = [
                f"{line.product.code} - {line.product.name}",
                line.lot_number or "",
                float(line.expected_quantity) if line.expected_quantity else None,
                float(line.received_quantity),
                line.unit.code if line.unit else "",
                float(line.unit_cost) if line.unit_cost else None,
                line.expiry_date.strftime("%d/%m/%Y") if line.expiry_date else "",
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_styles(cell, cs)

        if reception.observations:
            row += 2
            ws.cell(row=row, column=1, value="Observaciones:").font = label_font
            ws.cell(row=row, column=2, value=reception.observations)

        for col in range(1, len(line_headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="recepcion_{reception.code}.xlsx"'
        return response


# ---------------------------------------------------------------
# PDF — Listado de órdenes de compra
# ---------------------------------------------------------------

class PurchaseOrderListPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "procurement.view_purchaseorder"

    def get(self, request):
        orders = (
            PurchaseOrder.objects
            .select_related("supplier")
            .order_by("-order_date", "-id")
        )
        html_string = render_to_string(
            "procurement/pdf_order_list.html",
            {"orders": orders, "now": datetime.now()},
        )
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="ordenes_compra.pdf"'
        pisa.CreatePDF(io.StringIO(html_string), dest=response)
        return response


# ---------------------------------------------------------------
# Excel — Listado de órdenes de compra
# ---------------------------------------------------------------

class PurchaseOrderListExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "procurement.view_purchaseorder"

    def get(self, request):
        orders = (
            PurchaseOrder.objects
            .select_related("supplier")
            .order_by("-order_date", "-id")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Órdenes de Compra"

        headers = ["Número", "Proveedor", "RUC", "Fecha", "Estado", "Moneda", "Total"]
        hs = _header_style()
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            _apply_styles(cell, hs)

        cs = _cell_style()
        for row, oc in enumerate(orders, 2):
            values = [
                oc.number,
                oc.supplier.trade_name or oc.supplier.legal_name,
                oc.supplier.identification or "",
                oc.order_date.strftime("%d/%m/%Y") if oc.order_date else "",
                oc.get_status_display(),
                oc.currency,
                float(oc.total_amount),
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_styles(cell, cs)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="ordenes_compra.xlsx"'
        return response


# ---------------------------------------------------------------
# PDF — Detalle de una orden de compra
# ---------------------------------------------------------------

class PurchaseOrderDetailPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "procurement.view_purchaseorder"

    def get(self, request, pk):
        order = get_object_or_404(
            PurchaseOrder.objects.select_related("supplier", "created_by"), pk=pk
        )
        lines = order.lines.select_related("product", "unit").order_by("id")
        html_string = render_to_string(
            "procurement/pdf_order_detail.html",
            {"oc": order, "lines": lines, "now": datetime.now()},
        )
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="orden_{order.number}.pdf"'
        pisa.CreatePDF(io.StringIO(html_string), dest=response)
        return response


# ---------------------------------------------------------------
# Excel — Detalle de una orden de compra
# ---------------------------------------------------------------

class PurchaseOrderDetailExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "procurement.view_purchaseorder"

    def get(self, request, pk):
        order = get_object_or_404(
            PurchaseOrder.objects.select_related("supplier", "created_by"), pk=pk
        )
        lines = order.lines.select_related("product", "unit").order_by("id")

        wb = Workbook()
        ws = wb.active
        ws.title = "Orden de Compra"

        title_font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")
        ws["A1"].value = f"Orden de Compra {order.number}"
        ws["A1"].font = title_font

        label_font = Font(bold=True, size=10)
        info = [
            ("Proveedor:", order.supplier.trade_name or order.supplier.legal_name),
            ("RUC:", order.supplier.identification or ""),
            ("Fecha:", order.order_date.strftime("%d/%m/%Y") if order.order_date else ""),
            ("Estado:", order.get_status_display()),
            ("Moneda:", order.currency),
            ("Total:", str(order.total_amount)),
            ("Creado por:", str(order.created_by) if order.created_by else "—"),
        ]
        row = 3
        for label, value in info:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Detalle de líneas").font = Font(bold=True, size=12)
        row += 1

        line_headers = ["Producto", "Descripción", "Cantidad", "Unidad", "Precio Unit.", "Total Línea"]
        hs = _header_style()
        for col, h in enumerate(line_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            _apply_styles(cell, hs)

        cs = _cell_style()
        for line in lines:
            row += 1
            values = [
                f"{line.product.code} - {line.product.name}",
                line.description or "",
                float(line.quantity),
                line.unit.code if line.unit else "",
                float(line.unit_price),
                float(line.line_total),
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_styles(cell, cs)

        for col in range(1, len(line_headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="orden_{order.number}.xlsx"'
        return response
