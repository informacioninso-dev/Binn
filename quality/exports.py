# quality/exports.py
import io
from datetime import datetime

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from xhtml2pdf import pisa

from core.mixins import ModulePermissionMixin
from .models import (
    QAPlan, QualityInspection, ProductRecall,
    InspectionStage, LotStatus,
)


# ---------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------

def _thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=10),
        "fill": PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _thin_border(),
    }


def _cell_style():
    return {
        "font": Font(size=10),
        "alignment": Alignment(vertical="center", wrap_text=True),
        "border": _thin_border(),
    }


def _apply_styles(cell, styles):
    for attr, value in styles.items():
        setattr(cell, attr, value)


def _logo_path():
    return str(settings.BASE_DIR / "static" / "img" / "logo.png")


# ---------------------------------------------------------------
# PDF — Listado de inspecciones
# ---------------------------------------------------------------

class InspectionListPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "quality.view_qualityinspection"

    def get(self, request):
        inspections = (
            QualityInspection.objects
            .select_related("lot", "lot__product", "inspected_by", "plan")
            .order_by("-inspected_at")[:500]  # Limitar para PDF
        )
        html_string = render_to_string(
            "quality/pdf_inspection_list.html",
            {"inspections": inspections, "now": datetime.now(), "logo": _logo_path()},
        )
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="inspecciones_calidad.pdf"'
        pisa.CreatePDF(io.StringIO(html_string), dest=response)
        return response


# ---------------------------------------------------------------
# Excel — Listado de inspecciones
# ---------------------------------------------------------------

class InspectionListExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "quality.view_qualityinspection"

    def get(self, request):
        inspections = (
            QualityInspection.objects
            .select_related("lot", "lot__product", "inspected_by", "plan", "operation")
            .order_by("-inspected_at")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Inspecciones QA"

        headers = ["ID", "Lote", "Producto", "Código Prod.", "Etapa", "Resultado", "Fecha", "Inspector", "Plan QA"]
        hs = _header_style()
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            _apply_styles(cell, hs)

        cs = _cell_style()
        for row, insp in enumerate(inspections, 2):
            values = [
                insp.id,
                insp.lot.lot_number if insp.lot else "",
                insp.lot.product.name if insp.lot and insp.lot.product else "",
                insp.lot.product.code if insp.lot and insp.lot.product else "",
                insp.get_stage_display(),
                insp.get_result_display(),
                insp.inspected_at.strftime("%d/%m/%Y %H:%M") if insp.inspected_at else "",
                insp.inspected_by.get_full_name() if insp.inspected_by else "",
                insp.plan.name if insp.plan else "",
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_styles(cell, cs)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="inspecciones_calidad.xlsx"'
        return response


# ---------------------------------------------------------------
# PDF — Listado de planes de QA
# ---------------------------------------------------------------

class QAPlanListPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "quality.view_qaplan"

    def get(self, request):
        plans = (
            QAPlan.objects
            .select_related("product", "work_center", "route_step")
            .prefetch_related("parameters")
            .order_by("stage", "product__name", "name")
        )
        html_string = render_to_string(
            "quality/pdf_qa_plan_list.html",
            {"plans": plans, "now": datetime.now(), "logo": _logo_path()},
        )
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="planes_qa.pdf"'
        pisa.CreatePDF(io.StringIO(html_string), dest=response)
        return response


# ---------------------------------------------------------------
# Excel — Listado de planes de QA
# ---------------------------------------------------------------

class QAPlanListExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "quality.view_qaplan"

    def get(self, request):
        plans = (
            QAPlan.objects
            .select_related("product", "work_center", "route_step")
            .prefetch_related("parameters")
            .order_by("stage", "product__name", "name")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Planes QA"

        headers = ["Nombre", "Etapa", "Producto", "Centro Trabajo", "Paso Ruta", "Activo", "N° Parámetros"]
        hs = _header_style()
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            _apply_styles(cell, hs)

        cs = _cell_style()
        for row, plan in enumerate(plans, 2):
            values = [
                plan.name,
                plan.get_stage_display(),
                plan.product.name if plan.product else "Genérico",
                plan.work_center.name if plan.work_center else "",
                plan.route_step.name if plan.route_step else "",
                "Sí" if plan.is_active else "No",
                plan.parameters.count(),
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_styles(cell, cs)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="planes_qa.xlsx"'
        return response


# ---------------------------------------------------------------
# PDF — Listado de retiros de mercado
# ---------------------------------------------------------------

class RecallListPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "quality.view_productrecall"

    def get(self, request):
        recalls = (
            ProductRecall.objects
            .select_related("product", "origin")
            .prefetch_related("lots", "affected_clients")
            .order_by("-recall_date")
        )
        html_string = render_to_string(
            "quality/pdf_recall_list.html",
            {"recalls": recalls, "now": datetime.now(), "logo": _logo_path()},
        )
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="retiros_mercado.pdf"'
        pisa.CreatePDF(io.StringIO(html_string), dest=response)
        return response


# ---------------------------------------------------------------
# Excel — Listado de retiros de mercado
# ---------------------------------------------------------------

class RecallListExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "quality.view_productrecall"

    def get(self, request):
        recalls = (
            ProductRecall.objects
            .select_related("product", "origin")
            .prefetch_related("lots", "affected_clients")
            .order_by("-recall_date")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Retiros de Mercado"

        headers = ["Código", "Producto", "Origen", "Severidad", "Estado", "Fecha", "Lotes", "Clientes Afectados", "Motivo"]
        hs = _header_style()
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            _apply_styles(cell, hs)

        cs = _cell_style()
        for row, recall in enumerate(recalls, 2):
            values = [
                recall.code,
                recall.product.name if recall.product else "",
                recall.origin.name if recall.origin else "",
                recall.get_severity_display(),
                recall.get_status_display(),
                recall.recall_date.strftime("%d/%m/%Y") if recall.recall_date else "",
                recall.lots.count(),
                recall.affected_clients.count(),
                recall.reason[:100] if recall.reason else "",
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_styles(cell, cs)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="retiros_mercado.xlsx"'
        return response
