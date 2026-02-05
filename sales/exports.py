# sales/exports.py
import io
from datetime import datetime

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.views import View
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from xhtml2pdf import pisa

from core.mixins import ModulePermissionMixin
from .models import SaleOrder, SaleOrderLine


# ---------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------

def _thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=10),
        "fill": PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _thin_border(),
    }


def _cell_style():
    return {
        "font": Font(size=10),
        "alignment": Alignment(vertical="center", wrap_text=True),
        "border": _thin_border(),
    }


def _apply_styles(cell, styles):
    for attr, value in styles.items():
        setattr(cell, attr, value)


# ---------------------------------------------------------------
# PDF — Listado de pedidos de venta
# ---------------------------------------------------------------

class SaleOrderListPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "sales.view_saleorder"

    def get(self, request):
        orders = (
            SaleOrder.objects
            .select_related("client")
            .order_by("-created_at")
        )
        # Calculate totals
        for order in orders:
            order._total = order.get_total()

        html_string = render_to_string(
            "sales/pdf_order_list.html",
            {"orders": orders, "now": datetime.now()},
        )
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="pedidos_venta.pdf"'
        pisa.CreatePDF(io.StringIO(html_string), dest=response)
        return response


# ---------------------------------------------------------------
# Excel — Listado de pedidos de venta
# ---------------------------------------------------------------

class SaleOrderListExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "sales.view_saleorder"

    def get(self, request):
        orders = (
            SaleOrder.objects
            .select_related("client")
            .order_by("-created_at")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos de Venta"

        headers = ["Código", "Cliente", "RUC/CI", "Fecha Entrega", "Estado", "Método de Pago", "Total"]
        hs = _header_style()
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            _apply_styles(cell, hs)

        cs = _cell_style()
        for row, order in enumerate(orders, 2):
            values = [
                order.code,
                order.client.trade_name or order.client.legal_name,
                order.client.identification or "",
                order.delivery_date.strftime("%d/%m/%Y") if order.delivery_date else "",
                order.get_status_display(),
                order.payment_method or "",
                float(order.get_total()),
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_styles(cell, cs)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="pedidos_venta.xlsx"'
        return response


# ---------------------------------------------------------------
# PDF — Detalle de un pedido de venta
# ---------------------------------------------------------------

class SaleOrderDetailPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "sales.view_saleorder"

    def get(self, request, pk):
        order = get_object_or_404(
            SaleOrder.objects.select_related("client", "created_by"), pk=pk
        )
        lines = order.lines.select_related("product").order_by("id")
        total = order.get_total()
        html_string = render_to_string(
            "sales/pdf_order_detail.html",
            {"order": order, "lines": lines, "total": total, "now": datetime.now()},
        )
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="pedido_{order.code}.pdf"'
        pisa.CreatePDF(io.StringIO(html_string), dest=response)
        return response


# ---------------------------------------------------------------
# Excel — Detalle de un pedido de venta
# ---------------------------------------------------------------

class SaleOrderDetailExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "sales.view_saleorder"

    def get(self, request, pk):
        order = get_object_or_404(
            SaleOrder.objects.select_related("client", "created_by"), pk=pk
        )
        lines = order.lines.select_related("product").order_by("id")

        wb = Workbook()
        ws = wb.active
        ws.title = "Pedido de Venta"

        title_font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")
        ws["A1"].value = f"Pedido de Venta {order.code}"
        ws["A1"].font = title_font

        label_font = Font(bold=True, size=10)
        info = [
            ("Cliente:", order.client.trade_name or order.client.legal_name),
            ("RUC/CI:", order.client.identification or ""),
            ("Fecha entrega:", order.delivery_date.strftime("%d/%m/%Y") if order.delivery_date else ""),
            ("Estado:", order.get_status_display()),
            ("Método de pago:", order.payment_method or "—"),
            ("Total:", str(order.get_total())),
            ("Creado por:", str(order.created_by) if order.created_by else "—"),
        ]
        row = 3
        for label, value in info:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Detalle de líneas").font = Font(bold=True, size=12)
        row += 1

        line_headers = ["Producto", "Código", "Cantidad", "Precio Unit.", "Total Línea"]
        hs = _header_style()
        for col, h in enumerate(line_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            _apply_styles(cell, hs)

        cs = _cell_style()
        for line in lines:
            row += 1
            values = [
                line.product.name,
                line.product.code,
                float(line.quantity),
                float(line.unit_price),
                float(line.total_price),
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_styles(cell, cs)

        if order.notes:
            row += 2
            ws.cell(row=row, column=1, value="Notas:").font = label_font
            ws.cell(row=row, column=2, value=order.notes)

        for col in range(1, len(line_headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="pedido_{order.code}.xlsx"'
        return response
